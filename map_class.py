# 地图模块

from utils import *
import pickle
import numpy as np
import openpyxl

class MAP_CLASS:

    def __init__(self, path="maps/map002.xlsx"):

        self.Path = path

        wb_obj = openpyxl.load_workbook(path)

        # 读取表1基础地图
        #######################################################
        sheet_obj = wb_obj['map']

        self.XLength = sheet_obj.max_column
        self.YLength = sheet_obj.max_row
        self.MinX = 0
        self.MinY = 0
        self.MaxX = self.XLength - 1
        self.MaxY = self.YLength - 1

        self.ParkingSpaceMap = np.zeros((self.XLength, self.YLength))
        self.LaneMap = np.zeros((self.XLength, self.YLength))
        self.EntranceMap = np.zeros((self.XLength, self.YLength))
        self.ExitMap = np.zeros((self.XLength, self.YLength))
        self.EAEMap = np.zeros((self.XLength, self.YLength))
        self.ObstacleMap = np.zeros((self.XLength, self.YLength))
        self.ChargeStationMap = np.zeros((self.XLength, self.YLength))

        self.ObstacleList = []
        self.LaneList = []
        self.LaneUpList = []
        self.LaneDownList = []
        self.LaneLeftList = []
        self.LaneRightList = []

        num_cols = sheet_obj.max_column
        num_rows = sheet_obj.max_row

        for row in range(1, num_rows + 1):
            for col in range(1, num_cols + 1):
                cell_obj = sheet_obj.cell(row=row, column=col)
                if cell_obj.value:
                    contents = cell_obj.value.split(',')
                    for c in contents:

                        if c == 'P':
                            self.ParkingSpaceMap[col - 1, num_rows - row] = 1

                        elif c == 'L':
                            self.LaneMap[col - 1, num_rows - row] = 5
                            self.LaneList.append((col - 1, num_rows - row))
                        elif c == 'LU':
                            self.LaneMap[col - 1, num_rows - row] = 1
                            self.LaneUpList.append((col - 1, num_rows - row))
                        elif c == 'LD':
                            self.LaneMap[col - 1, num_rows - row] = 2
                            self.LaneDownList.append((col - 1, num_rows - row))
                        elif c == 'LL':
                            self.LaneMap[col - 1, num_rows - row] = 3
                            self.LaneLeftList.append((col - 1, num_rows - row))
                        elif c == 'LR':
                            self.LaneMap[col - 1, num_rows - row] = 4
                            self.LaneRightList.append((col - 1, num_rows - row))

                        elif c == 'I':
                            self.EntranceMap[col - 1, num_rows - row] = 1
                        elif c == 'O':
                            self.ExitMap[col - 1, num_rows - row] = 1
                        elif c == 'IO':
                            self.EAEMap[col - 1, num_rows - row] = 1

                        elif c == 'N':
                            self.ObstacleMap[col - 1, num_rows - row] = 1
                            self.ObstacleList.append((col - 1, num_rows - row))

                        elif c == 'C':
                            self.ChargeStationMap[col - 1, num_rows - row] = 1

        # 读取额外信息
        # 注意是直接读取坐标
        # 之所以需要额外开表是因为需要给出以下序号：出入口，AGV，AGV对应的充电站
        #######################################################

        self.EntranceList = []
        self.ExitList = []
        self.AGVInitPosList = []
        self.ChargeStationList = []

        # 入口
        ###########################

        sheet_obj = wb_obj['entrances']

        num_rows = sheet_obj.max_row

        for row in range(1, num_rows + 1):
            coord = []
            for col in range(1, 3):
                cell_obj = sheet_obj.cell(row=row, column=col)
                coord.append(cell_obj.value)
            self.EntranceList.append(tuple(coord))

        # 补齐到3个
        if len(self.EntranceList) < 3:
            for i in range(3 - len(self.EntranceList)):
                self.EntranceList.append((-1, -1))

        self.EntranceList = tuple(self.EntranceList)

        # 出口
        ###########################

        sheet_obj = wb_obj['exits']

        num_rows = sheet_obj.max_row

        for row in range(1, num_rows + 1):
            coord = []
            for col in range(1, 3):
                cell_obj = sheet_obj.cell(row=row, column=col)
                coord.append(cell_obj.value)
            self.ExitList.append(tuple(coord))

        if len(self.ExitList) < 3:
            for i in range(3 - len(self.ExitList)):
                self.ExitList.append((-1, -1))

        self.ExitList = tuple(self.ExitList)

        # agv初始位置和充电站
        ###########################

        sheet_obj = wb_obj['agvs']

        num_rows = sheet_obj.max_row

        for row in range(1, num_rows + 1):
            coord1 = []
            for col in range(1, 3):
                cell_obj = sheet_obj.cell(row=row, column=col)
                coord1.append(cell_obj.value)
            self.AGVInitPosList.append(tuple(coord1))
            coord2 = []
            for col in range(3, 5):
                cell_obj = sheet_obj.cell(row=row, column=col)
                coord2.append(cell_obj.value)
            self.ChargeStationList.append(tuple(coord1))

        self.AGVInitPosList = tuple(self.AGVInitPosList)
        self.ChargeStationList = tuple(self.ChargeStationList)

        # 读取条带信息
        #######################################################
        sheet_obj = wb_obj['stripes']

        num_rows = sheet_obj.max_row
        num_cols = sheet_obj.max_column

        # 条带信息列表
        self.StripeList = []
        for row in range(1, num_rows + 1):
            record = []
            for col in range(1, num_cols + 1):
                cell_obj = sheet_obj.cell(row=row, column=col)
                record.append(cell_obj.value)
            self.StripeList.append(tuple(record))

        self.StripeList = tuple(self.StripeList)
        self.NumStripes = len(self.StripeList)
        self.StripeDirectStepping = ((1, 0), (-1, 0), (0, 1), (0, -1))  # 方向定义

        # 读取pattern
        #######################################################
        sheet_obj = wb_obj['pattern']

        num_cols = sheet_obj.max_column

        pattern = []
        row = 1
        for col in range(1, num_cols + 1):
            cell_obj = sheet_obj.cell(row=row, column=col)
            pattern.append(cell_obj.value)
        self.PartitionPattern = tuple(pattern)  # 布局交线
        self.Dimension = len(self.PartitionPattern)  # 布局

    def size(self):
        return self.XLength, self.YLength

    def getEEInfo(self):
        return self.EntranceList, self.ExitList

    def getAGVInfo(self):
        return self.AGVInitPosList, self.ChargeStationList

    def getStripeInfo(self):
        stripe_length = []
        stripe_type = []
        for i in range(self.NumStripes):
            stripe_length.append(self.StripeList[i][3])
            stripe_type.append(self.StripeList[i][4])
        return stripe_length, stripe_type

    def getDimension(self):
        return self.Dimension

    def verifyPos(self, pos):
        if pos[0] < self.MinX:
            return False
        if pos[1] < self.MinY:
            return False
        if pos[0] > self.MaxX:
            return False
        if pos[1] > self.MaxY:
            return False
        #
        if self.ParkingSpaceMap[pos]:
            return False
        #
        if self.ObstacleMap[pos]:
            return False
        if self.ChargeStationMap[pos]:
            return False
        if self.EAEMap[pos]:
            return False
        if self.EntranceMap[pos]:
            return False
        if self.ExitMap[pos]:
            return False
        return True

    def genParkingPosListFromChromos(self, solu):

        # 由条带信息列表和解填充停车位表

        capacity = 0
        parking_post_list_A = []
        parking_post_list_B = []
        parking_post_list_C = []
        parking_post_list = []  # 总表

        # ①
        # for条带
        for index_stripe in range(self.NumStripes):
            # 读取条带信息
            start_x, start_y, direction, length, stripe_type = self.StripeList[index_stripe]
            # 累加容量
            capacity += length
            # 指针
            stepping = self.StripeDirectStepping[direction]
            pointer = (start_x, start_y)
            # 1分区式条带
            #######################################################
            if stripe_type == 0:
                # 从基因算法解中获取分区长度列表
                stripe_partition = solu[index_stripe]
                # ②b
                # for分区
                for index_part in range(self.Dimension):
                    # 获取分区类型
                    part_type = self.PartitionPattern[index_part]
                    # 获取分区长度
                    part_len = stripe_partition[index_part]
                    if part_type == 0:
                        li = parking_post_list_A
                    elif part_type == 1:
                        li = parking_post_list_B
                    elif part_type == 2:
                        li = parking_post_list_C
                    # ③
                    # for分区内pos
                    for l in range(part_len):
                        li.append((pointer[0],
                                   pointer[1],
                                   getDist(pointer, self.EntranceList[0]),
                                   getDist(pointer, self.EntranceList[1]),
                                   getDist(pointer, self.EntranceList[2]),
                                   getDist(pointer, self.ExitList[0]),
                                   getDist(pointer, self.ExitList[1]),
                                   getDist(pointer, self.ExitList[2]),
                                   0,
                                   -1
                                   ))
                        parking_post_list.append((pointer[0],
                                                  pointer[1],
                                                  getDist(pointer, self.EntranceList[0]),
                                                  getDist(pointer, self.EntranceList[1]),
                                                  getDist(pointer, self.EntranceList[2]),
                                                  getDist(pointer, self.ExitList[0]),
                                                  getDist(pointer, self.ExitList[1]),
                                                  getDist(pointer, self.ExitList[2]),
                                                  0,
                                                  -1
                                                  ))
                        # 分区内pos步进
                        pointer = (pointer[0] + stepping[0], pointer[1] + stepping[1])
            # 2非分区式条带
            #######################################################
            else:
                # 从基因算法解中获取条带马赛克
                stripe_mosaic = solu[index_stripe]
                # ②
                # for条带内pos
                for index_pos in range(length):
                    pos_type = stripe_mosaic[index_pos]
                    if pos_type == 0:
                        li = parking_post_list_A
                    elif pos_type == 1:
                        li = parking_post_list_B
                    elif pos_type == 2:
                        li = parking_post_list_C
                    li.append((pointer[0],
                               pointer[1],
                               getDist(pointer, self.EntranceList[0]),
                               getDist(pointer, self.EntranceList[1]),
                               getDist(pointer, self.EntranceList[2]),
                               getDist(pointer, self.ExitList[0]),
                               getDist(pointer, self.ExitList[1]),
                               getDist(pointer, self.ExitList[2]),
                               0,
                               -1
                               ))
                    parking_post_list.append((pointer[0],
                                              pointer[1],
                                              getDist(pointer, self.EntranceList[0]),
                                              getDist(pointer, self.EntranceList[1]),
                                              getDist(pointer, self.EntranceList[2]),
                                              getDist(pointer, self.ExitList[0]),
                                              getDist(pointer, self.ExitList[1]),
                                              getDist(pointer, self.ExitList[2]),
                                              0,
                                              -1
                                              ))
                    # 条带内pos步进
                    pointer = (pointer[0] + stepping[0], pointer[1] + stepping[1])

        return capacity, (parking_post_list_A, parking_post_list_B, parking_post_list_C), parking_post_list

    def getData(self):
        entrance_text_list = []
        exit_text_list = []
        charge_station_text_list = []
        for i in range(len(self.EntranceList)):
            entrance_text_list.append((self.EntranceList[i], '入' + str(i)))
        for i in range(len(self.ExitList)):
            exit_text_list.append((self.ExitList[i], '出' + str(i)))
        for i in range(len(self.ChargeStationList)):
            charge_station_text_list.append((self.ChargeStationList[i], '充' + str(i)))
        return entrance_text_list, exit_text_list, charge_station_text_list, self.ObstacleList, self.LaneList, self.LaneUpList, self.LaneDownList, self.LaneLeftList, self.LaneRightList


