import random
import pickle
import matplotlib.pyplot as plt
import math
import openpyxl


class TRAFFIC:
    """
    容量调整系数=地图容量/模型最大容量78.5(constant)
    上表到达概率单位为辆/小时，若以5秒为一step，则一小时有60*12个step
    所以每个step实际到达概率应该为a_adj*prob/(60*12)

    vehicle table = ['类型','分配类型','到达时间','离开时间','入口','出口','实际出口','是否进场','停车位id']
    """

    def __init__(self):
        self.PregeneratedVehicleList = []
        self.HourList = (0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24)

    def loadModel(self, path='models/model000.xlsx'):
        self.Path = path
        wb_obj = openpyxl.load_workbook(path)

        sheet_obj = wb_obj['short']
        num_rows = sheet_obj.max_row
        short = []
        for row in range(1, num_rows + 1):
            cell_obj = sheet_obj.cell(row=row, column=1)
            short.append(cell_obj.value)
        self.ShortParkProb = tuple(short)

        sheet_obj = wb_obj['middle']
        middle = []
        for row in range(1, num_rows + 1):
            cell_obj = sheet_obj.cell(row=row, column=1)
            middle.append(cell_obj.value)
        self.MiddleParkProb = tuple(middle)

        sheet_obj = wb_obj['long']
        long = []
        for row in range(1, num_rows + 1):
            cell_obj = sheet_obj.cell(row=row, column=1)
            long.append(cell_obj.value)
        self.LongParkProb = tuple(long)

        sheet_obj = wb_obj['definition']
        num_rows = sheet_obj.max_row
        defi = []
        for row in range(1, num_rows + 1):
            cell_obj = sheet_obj.cell(row=row, column=1)
            defi.append(cell_obj.value)
        self.ShortParkTill, self.MiddleParkTill, self.LongParkTill = defi

        sheet_obj = wb_obj['capacity']
        self.ModelCap = sheet_obj.cell(row=1, column=1).value
        self.TimeDiv = (self.ShortParkTill, self.MiddleParkTill, self.LongParkTill)

    def setParams(self, num_exp, map_cap, change_prob, entrance_percent, different_exit_prob):

        # 参数
        self.NumExp = num_exp
        self.MapCap = map_cap
        self.Ratio = map_cap / self.ModelCap
        self.ChangeProb = change_prob
        self.EntrancePercent = entrance_percent
        self.DifferentExitProb = different_exit_prob

    def entranceChoice(self):
        a = self.EntrancePercent[0]
        b = a + self.EntrancePercent[1]
        # c = b + self.EntrancePercent[2]
        r = random.random()
        if r < a:
            return 0
        elif r < b:
            return 1
        else:
            return 2

    def genVehicleList(self):

        vehicle_table = []

        for step in range(24 * 60 * 12):
            # 插值求概率
            hour = step / (60 * 12)
            if not int(hour) == hour:
                x1 = math.floor(hour)
                x2 = math.ceil(hour)

                y1 = self.ShortParkProb[x1]
                y2 = self.ShortParkProb[x2]
                probA = (y2 - y1) / 1 * (hour - x1) + y1

                y1 = self.MiddleParkProb[x1]
                y2 = self.MiddleParkProb[x2]
                probB = (y2 - y1) / 1 * (hour - x1) + y1

                y1 = self.LongParkProb[x1]
                y2 = self.LongParkProb[x2]
                probC = (y2 - y1) / 1 * (hour - x1) + y1
            else:
                probA = self.ShortParkProb[int(hour)]
                probB = self.MiddleParkProb[int(hour)]
                probC = self.LongParkProb[int(hour)]
            # 采样
            # A采样
            if random.random() < probA * self.Ratio / (60 * 12):
                if random.random() < self.ChangeProb:
                    leave_after = random.uniform(self.ShortParkTill, self.LongParkTill)
                else:
                    leave_after = random.uniform(0.2, self.ShortParkTill)  #
                leave_step = int((hour + leave_after) * 60 * 12)
                # 修正离开时间
                while leave_step > 17279:
                    # print('A'+str(leave_step))
                    leave_after = random.uniform(0.2, self.ShortParkTill)
                    leave_step = int((hour + leave_after) * 60 * 12)
                # 入口抽样
                ent = self.entranceChoice()
                # ['类型','分配类型','到达时间','离开时间','入口','出口','实际出口','是否进场','停车位id']
                vehicle_table.append((1, -1, step, leave_step, ent, ent, ent, 0, -1))
            # B采样
            if random.random() < probB * self.Ratio / (60 * 12):
                if random.random() < self.ChangeProb:
                    if random.random() < 0.5:
                        leave_after = random.uniform(0.2, self.ShortParkTill)
                    else:
                        leave_after = random.uniform(self.MiddleParkTill, self.LongParkTill)
                else:
                    leave_after = random.uniform(self.ShortParkTill, self.MiddleParkTill)  #
                leave_step = int((hour + leave_after) * 60 * 12)
                #
                while leave_step > 17279:
                    # print('B'+str(leave_step))
                    leave_after = random.uniform(self.ShortParkTill, self.MiddleParkTill)
                    leave_step = int((hour + leave_after) * 60 * 12)
                    # 入口抽样
                ent = self.entranceChoice()
                vehicle_table.append((2, -1, step, leave_step, ent, ent, ent, 0, -1))
            # C采样
            if random.random() < probC * self.Ratio / (60 * 12):
                if random.random() < self.ChangeProb:
                    leave_after = random.uniform(0.2, self.MiddleParkTill)
                else:
                    leave_after = random.uniform(self.MiddleParkTill, self.LongParkTill)  #
                leave_step = int((hour + leave_after) * 60 * 12)
                #
                while leave_step > 17279:
                    # print('C'+str(leave_step))
                    leave_after = random.uniform(self.MiddleParkTill, self.LongParkTill)
                    leave_step = int((hour + leave_after) * 60 * 12)
                    # 入口抽样
                ent = self.entranceChoice()
                vehicle_table.append((3, -1, step, leave_step, ent, ent, ent, 0, -1))

        arrive_queue = []
        for i in range(len(vehicle_table)):
            arrive_queue.append((vehicle_table[i][2], i))  # 到达时间，id

        return vehicle_table, arrive_queue

    def generate(self):
        for i in range(self.NumExp):
            self.PregeneratedVehicleList.append(self.genVehicleList())

    def info(self):
        print('实验数量：' + str(self.NumExp))
        print('地图容量：' + str(self.MapCap))
        print('模型容量：' + str(self.ModelCap))
        print('模型路径：' + str(self.Path))
        print('离开变类概率：' + str(self.ChangeProb))
        print('入口百分比：' + str(self.EntrancePercent))
        print('异口离开概率：' + str(self.DifferentExitProb))
        print('实际时间分割：' + str(self.TimeDiv))

    def save(self):
        print('实验数量：' + str(self.NumExp))
        print('地图容量：' + str(self.MapCap))
        print('模型容量：' + str(self.ModelCap))
        print('模型路径：' + str(self.Path))
        print('离开变类概率：' + str(self.ChangeProb))
        print('入口百分比：' + str(self.EntrancePercent))
        print('异口离开概率：' + str(self.DifferentExitProb))
        print('实际时间分割：' + str(self.TimeDiv))
        file_name = input("请输入文件名，输入exit取消：")
        if file_name == 'exit':
            return
        with open(file_name + '.bin', 'wb') as f:
            pickle.dump((self.PregeneratedVehicleList, (
                self.NumExp, self.MapCap, self.Path, self.ModelCap, self.ChangeProb, self.EntrancePercent,
                self.DifferentExitProb, self.TimeDiv)), f)

    def load(self):
        file_name = input("请输入文件名，输入exit取消：")
        if file_name == 'exit':
            return
        with open(file_name + '.bin', 'rb') as f:
            self.PregeneratedVehicleList, _ = pickle.load(f)
            self.NumExp, self.MapCap, self.Path, self.ModelCap, self.ChangeProb, self.EntrancePercent, self.DifferentExitProb, self.TimeDiv = _
            print('已成功加载文件：' + str(file_name) + '.bin')
            print('实验数量：' + str(self.NumExp))
            print('地图容量：' + str(self.MapCap))
            print('模型容量：' + str(self.ModelCap))
            print('模型路径：' + str(self.Path))
            print('离开变类概率：' + str(self.ChangeProb))
            print('入口百分比：' + str(self.EntrancePercent))
            print('异口离开概率：' + str(self.DifferentExitProb))
            print('实际时间分割：' + str(self.TimeDiv))

    def plotModel(self):
        fig, ax = plt.subplots()
        ax.plot(self.HourList, self.ShortParkProb, label='short')
        ax.plot(self.HourList, self.MiddleParkProb, label='middle')
        ax.plot(self.HourList, self.LongParkProb, label='long')
        ax.legend()
        plt.show()

    def reclassify(self, time_div):
        self.TimeDiv = tuple(time_div)
        time_div = [e * 60 * 12 for e in time_div]
        for i in range(self.NumExp):
            exp_vehicle_table = self.PregeneratedVehicleList[i][0]
            for ii in range(len(exp_vehicle_table)):
                record = list(exp_vehicle_table[ii])
                enter_step = record[2]
                leave_step = record[3]
                park_duration = leave_step - enter_step
                new_category = next(i for i, v in enumerate(time_div) if v > park_duration) + 1
                record[0] = new_category
                exp_vehicle_table[ii] = tuple(record)

    def addClassificationFailure(self, prob):
        for i in range(self.NumExp):
            exp_vehicle_table = self.PregeneratedVehicleList[i][0]
            for ii in range(len(exp_vehicle_table)):
                if random.random() < prob:
                    record = list(exp_vehicle_table[ii])
                    enter_step = record[2]
                    leave_step = record[3]
                    lower_bound = max((leave_step - 4 * 60 * 12), (enter_step + 10 * 12))
                    upper_bound = leave_step + 5 * 60 * 12
                    leave_step = int(random.uniform(lower_bound, upper_bound))
                    record[3] = leave_step
                    exp_vehicle_table[ii] = tuple(record)
        self.ChangeProb = prob
